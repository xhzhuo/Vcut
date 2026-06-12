"""Integration tests for manual xlsx-driven pipeline mode."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from vcut.core.pipeline import run_pipeline
from vcut.core.pipeline_manual import run_manual_pipeline

ZH_INDEX = "\u5e8f\u53f7"
ZH_VIDEO = "\u89c6\u9891"
ZH_PAIN = "\u75db\u70b9"
ZH_SCENE = "\u4f7f\u7528\u573a\u666f"
ZH_BENEFIT = "\u6210\u5206\u529f\u6548"
ZH_CTA = "\u673a\u5236\u53f7\u53ec"


def _fake_transcript_index(**_kwargs) -> dict:
    return {}


def _fake_plan_builder_for_labels(video_dir: Path):
    def _fake_build_manual_edit_plans(
        segments,
        labels,
        variants,
        max_total_duration=None,
        use_llm=False,
        **_kwargs,
    ):
        assert use_llm is True
        selected = []
        for label in labels:
            match = next(segment for segment in segments if segment["label"] == label)
            selected.append(
                {
                    "segment_id": match["segment_id"],
                    "video_id": Path(match["src_video"]).name,
                    "src_video": str(Path(match["src_video"]).resolve(strict=False)),
                    "start": match["start"],
                    "end": match["end"],
                    "duration": match["duration"],
                    "reason": "llm-selected",
                    "score": 1.0,
                    "role": "clip",
                }
            )
        if selected:
            selected[0]["role"] = "hook"
            selected[-1]["role"] = "closing"
        return [selected for _ in range(max(1, int(variants)))]

    return _fake_build_manual_edit_plans


def test_run_pipeline_manual_mode_writes_artifacts_and_calls_render(monkeypatch, tmp_path) -> None:
    root = tmp_path / "manual_pipeline"
    root.mkdir(parents=True, exist_ok=True)
    video_dir = root / "videos"
    video_dir.mkdir(parents=True, exist_ok=True)
    (video_dir / "1.mp4").write_bytes(b"a")
    (video_dir / "2.mp4").write_bytes(b"b")

    xlsx_path = root / "plan.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([ZH_INDEX, ZH_VIDEO, ZH_PAIN, ZH_SCENE, ZH_BENEFIT, ZH_CTA])
    ws.append([1, "1.mp4", "0s-5s", "5s-8s", "8s-12s", "12s-15s"])
    ws.append([2, "2.mp4", "1s-4s", "4s-7s", "7s-10s", "10s-13s"])
    wb.save(xlsx_path)

    artifacts_dir = root / "artifacts"
    config = {
        "artifacts_dir": str(artifacts_dir),
        "render": {"enabled": True},
    }
    called = {"render": 0}

    def fake_run_ffmpeg(command):
        called["render"] += 1
        output_path = Path(command[-1])
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(b"out")

    monkeypatch.setattr("vcut.stages.video_edit._run_ffmpeg", fake_run_ffmpeg)

    run_manual_pipeline(
        config=config,
        artifacts_dir=artifacts_dir,
        output_video=str(root / "out.mp4"),
        manual_xlsx=str(xlsx_path),
        manual_video_dir=str(video_dir),
        manual_labels=[ZH_PAIN, ZH_SCENE, ZH_BENEFIT, ZH_CTA],
        manual_variants=1,
        manual_max_duration=None,
        manual_use_asr_llm=True,
        manual_goal=None,
        build_transcript_index_fn=_fake_transcript_index,
        build_manual_edit_plans_fn=_fake_plan_builder_for_labels(video_dir),
    )

    assert called["render"] == 5  # 4 cuts + 1 concat
    assert (artifacts_dir / "manual_segments.json").exists()
    # Edit plan is now named per-output: edit_plan_{stem}.json
    edit_plan_files = list(artifacts_dir.glob("edit_plan_*.json"))
    assert len(edit_plan_files) >= 1
    plan = json.loads(edit_plan_files[0].read_text(encoding="utf-8"))
    assert len(plan) == 4
    assert plan[0]["role"] == "hook"
    assert plan[-1]["role"] == "closing"


def test_run_pipeline_manual_mode_asr_llm_uses_full_segments(monkeypatch, tmp_path) -> None:
    root = tmp_path / "manual_pipeline"
    root.mkdir(parents=True, exist_ok=True)
    video_dir = root / "videos"
    video_dir.mkdir(parents=True, exist_ok=True)
    (video_dir / "1.mp4").write_bytes(b"a")
    (video_dir / "2.mp4").write_bytes(b"b")

    xlsx_path = root / "plan.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([ZH_INDEX, ZH_VIDEO, ZH_PAIN, ZH_SCENE])
    ws.append([1, "1.mp4", "0s-5s", "5s-8s"])
    ws.append([2, "2.mp4", "1s-4s", "4s-7s"])
    wb.save(xlsx_path)

    artifacts_dir = root / "artifacts"
    config = {
        "artifacts_dir": str(artifacts_dir),
        "render": {"enabled": False},
        "asr": {"model_name": "base"},
        "strategy": {"model_name": "x"},
    }
    captured = {"asr_segments_count": 0, "llm_used": False}

    monkeypatch.setattr("vcut.core.pipeline.load_config", lambda _: config)
    monkeypatch.setenv("DOUBAO_ASR_API_KEY", "test-key")

    def fake_transcribe_to_artifacts(video_path, transcript_json_path, transcript_srt_path, model_name="base", model=None, asr_options=None, asr_config=None):
        captured["asr_segments_count"] = captured.get("asr_segments_count", 0) + 1
        return {"segments": [{"start": 0.0, "end": 8.0, "text": "hello"}], "text": "hello", "provider": "test"}

    def fake_build_manual_edit_plans(
        segments,
        labels,
        variants,
        max_total_duration=None,
        use_llm=False,
        llm_goal=None,
        llm_model_name=None,
        llm_api_key_env=None,
        llm_endpoint=None,
        unique_src_video=False,
        variant_offset=0,
        prior_plans=None,
        used_combinations=None,
        quality_config=None,
        review_config=None,
        review_log=None,
    ):
        captured["llm_used"] = use_llm
        return [
            [
                {
                    "segment_id": "x1",
                    "video_id": "1.mp4",
                    "src_video": str((video_dir / "1.mp4").resolve()),
                    "start": 0.0,
                    "end": 5.0,
                    "duration": 5.0,
                    "reason": "x",
                    "score": 1.0,
                    "role": "hook",
                }
            ]
        ]

    monkeypatch.setattr("vcut.stages.asr.transcribe_to_artifacts", fake_transcribe_to_artifacts)
    monkeypatch.setattr("vcut.manual.asr.transcribe_to_artifacts", fake_transcribe_to_artifacts)
    monkeypatch.setenv("OPENAI_API_KEY", "dummy")
    monkeypatch.setattr(
        "vcut.manual.strategy._select_with_llm",
        lambda segments, labels, **kwargs: [
            {"segment_id": s["segment_id"], "label": l, "reason": "ok"}
            for s, l in zip(segments[:len(labels)], labels)
        ],
    )

    run_pipeline(
        output_video=str(root / "out.mp4"),
        manual_xlsx=str(xlsx_path),
        manual_video_dir=str(video_dir),
        manual_labels=[ZH_PAIN, ZH_SCENE],
        manual_variants=1,
        manual_use_asr_llm=True,
    )

    assert captured["asr_segments_count"] >= 1
    assert (artifacts_dir / "manual_segments.json").exists()


def test_run_pipeline_applies_manual_review_criteria_override(monkeypatch, tmp_path) -> None:
    captured = {}
    config = {"artifacts_dir": str(tmp_path / "artifacts"), "strategy": {"review": {}}}

    monkeypatch.setattr("vcut.core.pipeline.load_config", lambda _: config)

    def fake_run_manual_pipeline(**kwargs):
        captured["config"] = kwargs["config"]

    monkeypatch.setattr("vcut.core.pipeline.run_manual_pipeline", fake_run_manual_pipeline)

    run_pipeline(
        output_video=str(tmp_path / "out.mp4"),
        manual_xlsx=str(tmp_path / "plan.xlsx"),
        manual_video_dir=str(tmp_path),
        manual_labels=[ZH_PAIN],
        manual_review_criteria="必须检查结尾是否自然收束。",
    )

    review_config = captured["config"]["strategy"]["review"]
    assert review_config["criteria"] == "必须检查结尾是否自然收束。"


def test_run_pipeline_manual_mode_records_signature_after_render(monkeypatch, tmp_path) -> None:
    root = tmp_path / "manual_pipeline_signature"
    root.mkdir(parents=True, exist_ok=True)
    video_dir = root / "videos"
    video_dir.mkdir(parents=True, exist_ok=True)
    (video_dir / "1.mp4").write_bytes(b"a")

    xlsx_path = root / "plan.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([ZH_INDEX, ZH_VIDEO, ZH_PAIN])
    ws.append([1, "1.mp4", "0s-3s"])
    wb.save(xlsx_path)

    artifacts_dir = root / "artifacts"
    config = {
        "artifacts_dir": str(artifacts_dir),
        "render": {"enabled": True},
    }

    def fake_run_ffmpeg(command):
        output_path = Path(command[-1])
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_bytes(b"out")

    monkeypatch.setattr("vcut.stages.video_edit._run_ffmpeg", fake_run_ffmpeg)

    run_manual_pipeline(
        config=config,
        artifacts_dir=artifacts_dir,
        output_video=str(root / "out.mp4"),
        manual_xlsx=str(xlsx_path),
        manual_video_dir=str(video_dir),
        manual_labels=[ZH_PAIN],
        manual_variants=1,
        manual_max_duration=None,
        manual_use_asr_llm=True,
        manual_goal=None,
        build_transcript_index_fn=_fake_transcript_index,
        build_manual_edit_plans_fn=_fake_plan_builder_for_labels(video_dir),
    )

    history_path = artifacts_dir / "used_plan_signatures.txt"
    assert history_path.exists()
    assert "1_痛点_001" in history_path.read_text(encoding="utf-8")


def test_run_pipeline_manual_mode_writes_rejected_reviews_before_failure(tmp_path) -> None:
    root = tmp_path / "manual_pipeline_rejected_reviews"
    root.mkdir(parents=True, exist_ok=True)
    video_dir = root / "videos"
    video_dir.mkdir(parents=True, exist_ok=True)
    (video_dir / "1.mp4").write_bytes(b"a")

    xlsx_path = root / "plan.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([ZH_INDEX, ZH_VIDEO, ZH_PAIN])
    ws.append([1, "1.mp4", "0s-3s"])
    wb.save(xlsx_path)

    artifacts_dir = root / "artifacts"

    def fake_build_manual_edit_plans(
        segments,
        labels,
        variants,
        max_total_duration=None,
        use_llm=False,
        llm_goal=None,
        llm_model_name=None,
        llm_api_key_env=None,
        llm_endpoint=None,
        unique_src_video=False,
        variant_offset=0,
        prior_plans=None,
        used_combinations=None,
        quality_config=None,
        review_config=None,
        review_log=None,
    ):
        if review_log is not None:
            review_log.append(
                {
                    "status": "rejected",
                    "variant_index": variant_offset,
                    "selected_segment_ids": ["bad"],
                    "edit_plan": [],
                    "review": {"approved": False, "score": 40, "issues": ["bad bridge"]},
                }
            )
        return []

    with pytest.raises(RuntimeError):
        run_manual_pipeline(
            config={"strategy": {"review": {"enabled": True}}, "render": {"enabled": False}},
            artifacts_dir=artifacts_dir,
            output_video=str(root / "out.mp4"),
            manual_xlsx=str(xlsx_path),
            manual_video_dir=str(video_dir),
            manual_labels=[ZH_PAIN],
            manual_variants=1,
            manual_max_duration=None,
            manual_use_asr_llm=True,
            manual_goal=None,
            build_transcript_index_fn=_fake_transcript_index,
            build_manual_edit_plans_fn=fake_build_manual_edit_plans,
        )

    rejected_path = artifacts_dir / "rejected_plans.jsonl"
    assert rejected_path.exists()
    assert "bad bridge" in rejected_path.read_text(encoding="utf-8")
def test_run_pipeline_manual_mode_groups_outputs_by_input_folder(monkeypatch, tmp_path) -> None:
    workspace_root = tmp_path / "workspace"
    group_name = "test_brand_manual"
    group_dir = workspace_root / "inputs" / group_name
    group_dir.mkdir(parents=True, exist_ok=True)
    video_dir = group_dir
    (video_dir / "1.mp4").write_bytes(b"a")

    xlsx_path = group_dir / "鍒囩墖鏂规.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append([ZH_INDEX, ZH_VIDEO, ZH_PAIN])
    ws.append([1, "1.mp4", "0s-3s"])
    wb.save(xlsx_path)

    artifacts_dir = tmp_path / "artifacts_manual_grouped"

    config = {
        "artifacts_dir": str(artifacts_dir),
        "render": {"enabled": True},
    }
    captured: dict[str, str] = {}

    monkeypatch.setattr("vcut.core.pipeline.load_config", lambda _: config)
    monkeypatch.setattr("vcut.core.pipeline_paths.workspace_root", lambda: workspace_root)

    def fake_run_manual_pipeline(**kwargs):
        captured["output_video"] = kwargs["output_video"]
        kwargs["artifacts_dir"].mkdir(parents=True, exist_ok=True)
        (kwargs["artifacts_dir"] / "manual_segments.json").write_text("[]", encoding="utf-8")
        (kwargs["artifacts_dir"] / "edit_plan_out.json").write_text("[]", encoding="utf-8")

    monkeypatch.setattr("vcut.core.pipeline.run_manual_pipeline", fake_run_manual_pipeline)

    run_pipeline(
        output_video="artifacts/out.mp4",
        manual_xlsx=str(xlsx_path),
        manual_video_dir=str(video_dir),
        manual_labels=[ZH_PAIN],
        manual_variants=1,
        manual_use_asr_llm=True,
    )

    grouped_dir = artifacts_dir / group_name
    assert grouped_dir.exists()
    assert (grouped_dir / "manual_segments.json").exists()
    edit_plan_files = list(grouped_dir.glob("edit_plan_*.json"))
    assert len(edit_plan_files) >= 1
    assert Path(captured["output_video"]) == (grouped_dir / "out.mp4").resolve(strict=False)

