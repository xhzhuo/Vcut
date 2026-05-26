"""Utilities for parsing manual segment annotations from Excel."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook

from vcut.io.fingerprint import canonical_path

_TIME_RANGE_SPLIT = re.compile(r"\s*[-~锝炩€斺€擄紞]\s*")
_TIME_TOKEN_CLEAN = re.compile(r"[^0-9:.]")
_PLAIN_FLOAT_RE = re.compile(r"^\d+(?:\.\d+)?$")
_COLON_PAIR_RE = re.compile(r"^(\d+):(\d+(?:\.\d+)?)$")
_HEADER_KEY_CLEAN = re.compile(r"[\s_/\-]+")

ZH_VIDEO = "\u89c6\u9891"
ZH_INDEX = "\u5e8f\u53f7"
ZH_PAIN = "\u75db\u70b9"
ZH_SCENE = "\u4f7f\u7528\u573a\u666f"
ZH_SCENE_ALT = "\u9002\u7528\u75c7\u72b6/\u573a\u666f"
ZH_SCENE_ALT_FLAT = "\u9002\u7528\u75c7\u72b6\u573a\u666f"
ZH_SCENE_SHORT = "\u573a\u666f"
ZH_BENEFIT = "\u6210\u5206\u529f\u6548"
ZH_CTA = "\u673a\u5236\u53f7\u53ec"

_HEADER_ALIASES = {
    ZH_VIDEO: (ZH_VIDEO, "video", "videofile", "sourcevideo"),
    ZH_INDEX: (ZH_INDEX, "\u7f16\u53f7", "index", "id"),
}

_LABEL_ALIASES = {
    ZH_PAIN: (ZH_PAIN,),
    ZH_SCENE: (ZH_SCENE, ZH_SCENE_ALT, ZH_SCENE_ALT_FLAT, ZH_SCENE_SHORT),
    ZH_BENEFIT: (ZH_BENEFIT,),
    ZH_CTA: (ZH_CTA,),
}


def _compact_key(value: str) -> str:
    return _HEADER_KEY_CLEAN.sub("", str(value or "").strip().lower())


def normalize_manual_label(label: str) -> str:
    """Normalize label aliases to stable canonical labels."""
    raw = str(label or "").strip()
    if not raw:
        return ""
    key = _compact_key(raw)
    for canonical, aliases in _LABEL_ALIASES.items():
        if key in {_compact_key(alias) for alias in aliases}:
            return canonical
    return raw


def _normalize_header(value: str) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    key = _compact_key(raw)
    for canonical, aliases in _HEADER_ALIASES.items():
        if key in {_compact_key(alias) for alias in aliases}:
            return canonical
    return raw


def _parse_time_token(token: str, *, frame_rate: float, sec_frame_mode: bool) -> float:
    cleaned = _TIME_TOKEN_CLEAN.sub("", str(token or "").strip().lower().replace("sec", "s"))
    if not cleaned:
        raise ValueError(f"Invalid time token: {token!r}")

    if cleaned.endswith("s"):
        number_part = cleaned[:-1]
        if _PLAIN_FLOAT_RE.match(number_part):
            return float(number_part)

    if _PLAIN_FLOAT_RE.match(cleaned):
        return float(cleaned)

    colon_count = cleaned.count(":")
    if colon_count == 2:
        first_str, second_str, third_str = cleaned.split(":", maxsplit=2)
        first = float(first_str)
        second = float(second_str)
        third = float(third_str)
        if sec_frame_mode:
            return (first * 60.0) + second + (third / frame_rate)
        return (first * 3600.0) + (second * 60.0) + third

    if colon_count == 1:
        match = _COLON_PAIR_RE.match(cleaned)
        if not match:
            raise ValueError(f"Invalid time token: {token!r}")
        left = float(match.group(1))
        right = float(match.group(2))
        if sec_frame_mode:
            return left + (right / frame_rate)
        return (left * 60.0) + right

    raise ValueError(f"Invalid time token: {token!r}")


def _infer_sec_frame_mode(chunks: list[str], frame_rate: float) -> bool:
    pair_values: list[float] = []
    triple_last_values: list[float] = []
    for chunk in chunks:
        parts = [part.strip() for part in _TIME_RANGE_SPLIT.split(chunk, maxsplit=1)]
        if len(parts) != 2:
            continue
        for part in parts:
            cleaned = _TIME_TOKEN_CLEAN.sub("", part.lower().replace("sec", "s"))
            if cleaned.count(":") != 1:
                if cleaned.count(":") == 2:
                    triple = cleaned.split(":", maxsplit=2)
                    try:
                        triple_last_values.append(float(triple[2]))
                    except ValueError:
                        pass
                continue
            match = _COLON_PAIR_RE.match(cleaned)
            if not match:
                continue
            pair_values.append(float(match.group(2)))
    pair_mode = bool(pair_values) and all(value < frame_rate for value in pair_values)
    triple_mode = bool(triple_last_values) and all(value < frame_rate for value in triple_last_values)
    return pair_mode or triple_mode


def parse_time_ranges(text: str, *, frame_rate: float = 30.0) -> list[tuple[float, float]]:
    """Parse manual time ranges into second ranges."""
    if not str(text or "").strip():
        return []
    if frame_rate <= 0:
        raise ValueError("frame_rate must be > 0.")

    normalized = (
        str(text)
        .replace("\uff0c", "/")
        .replace("\u3001", "/")
        .replace("\uff1b", "/")
        .replace(";", "/")
        .replace(",", "/")
    )
    chunks = [chunk.strip() for chunk in normalized.split("/") if chunk and chunk.strip()]
    sec_frame_mode = _infer_sec_frame_mode(chunks, frame_rate)

    ranges: list[tuple[float, float]] = []
    for chunk in chunks:
        parts = [part.strip() for part in _TIME_RANGE_SPLIT.split(chunk, maxsplit=1)]
        if len(parts) != 2:
            raise ValueError(f"Invalid time range: {text!r}")
        start = _parse_time_token(parts[0], frame_rate=frame_rate, sec_frame_mode=sec_frame_mode)
        end = _parse_time_token(parts[1], frame_rate=frame_rate, sec_frame_mode=sec_frame_mode)
        if end <= start:
            raise ValueError(f"Invalid time range with end<=start: {text!r}")
        ranges.append((round(start, 3), round(end, 3)))
    return ranges


def _normalize_headers(header_row: tuple) -> list[str]:
    return [_normalize_header(str(cell).strip()) if cell is not None else "" for cell in header_row]


def load_manual_segments_from_excel(
    xlsx_path: str,
    video_dir: str,
    *,
    sheet_name: str | None = None,
    video_column: str = ZH_VIDEO,
    skip_columns: tuple[str, ...] = (ZH_INDEX,),
    timecode_fps: float = 30.0,
) -> list[dict]:
    """Load manually labeled segments from xlsx file."""
    xlsx = Path(xlsx_path)
    if not xlsx.exists() or not xlsx.is_file():
        raise FileNotFoundError(f"Manual xlsx not found: {xlsx_path}")

    root = Path(video_dir)
    if not root.exists() or not root.is_dir():
        raise FileNotFoundError(f"Manual video dir not found: {video_dir}")

    workbook = load_workbook(xlsx, data_only=True)
    sheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    if sheet.max_row < 2:
        return []

    headers = _normalize_headers(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    normalized_video_column = _normalize_header(video_column)
    normalized_skip_columns = {_normalize_header(item) for item in skip_columns}

    if normalized_video_column not in headers:
        raise ValueError(f"Missing required column `{video_column}` in {xlsx_path}")
    video_idx = headers.index(normalized_video_column)

    label_indices: list[tuple[str, int]] = []
    for idx, header in enumerate(headers):
        if not header:
            continue
        if header == normalized_video_column or header in normalized_skip_columns:
            continue
        label_indices.append((header, idx))

    if not label_indices:
        raise ValueError(f"No label columns found in {xlsx_path}")

    segments: list[dict] = []
    per_label_counter: dict[tuple[str, str], int] = {}
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
        video_file = str(row[video_idx] or "").strip()
        if not video_file:
            continue

        src_video_path = Path(video_file)
        if not src_video_path.is_absolute():
            src_video_path = root / video_file
        src_video = canonical_path(str(src_video_path))
        if not Path(src_video).exists():
            raise FileNotFoundError(f"Source video from xlsx row not found: {src_video}")

        video_key = Path(video_file).stem.replace(" ", "_")
        for label, idx in label_indices:
            cell_value = str(row[idx] or "").strip()
            if not cell_value:
                continue
            for start, end in parse_time_ranges(cell_value, frame_rate=timecode_fps):
                key = (video_file, label)
                per_label_counter[key] = per_label_counter.get(key, 0) + 1
                seg_no = per_label_counter[key]
                segments.append(
                    {
                        "segment_id": f"{video_key}_{label}_{seg_no:03d}",
                        "video_file": video_file,
                        "src_video": src_video,
                        "label": label,
                        "start": start,
                        "end": end,
                        "duration": round(end - start, 3),
                    }
                )

    if not segments:
        raise ValueError(f"No manual segments parsed from {xlsx_path}")
    return segments


def write_manual_segments_json(segments: list[dict], output_path: str | Path) -> None:
    """Write parsed manual segments to json."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(segments, ensure_ascii=False, indent=2), encoding="utf-8")

